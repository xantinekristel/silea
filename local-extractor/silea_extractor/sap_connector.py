"""SAP GUI Scripting connector (Phase 3) — LOCAL WINDOWS ONLY.

Attaches to an already-running, already-logged-in SAP GUI session and runs
ZRGCSTAT with Role Category = "Buyers Only", then exports the result. It never
launches SAP, never logs in, and never stores a password — the staff member is
already authenticated (SPEC §10, non-functional requirement "No credentials in
code").

This is a SCAFFOLD. It cannot run or be tested in a cloud sandbox: it needs
Windows, the SAP GUI client, and SAP GUI Scripting enabled by IT. The
transaction/field IDs below are placeholders to be confirmed against the real
ZRGCSTAT screen during on-site wiring. Do not begin this wiring until IT has
answered the questions in README.md / SPEC §10.
"""

from __future__ import annotations

import os
from typing import List

from .models import Account


class LiveSapConnector:
    name = "SAP GUI Scripting (ZRGCSTAT)"
    live = True

    def __init__(self, export_dir: str) -> None:
        self.export_dir = export_dir

    def _session(self):
        # Imported lazily so the module imports fine off-Windows.
        try:
            import win32com.client  # type: ignore  # from pywin32
        except ImportError as exc:  # pragma: no cover - platform specific
            raise RuntimeError(
                "pywin32 is required for the live SAP connector (Windows only). "
                "Install with: pip install pywin32"
            ) from exc

        sap_gui = win32com.client.GetObject("SAPGUI")
        app = sap_gui.GetScriptingEngine
        if app.Connections.Count == 0:  # pragma: no cover
            raise RuntimeError(
                "No open SAP connection. Log in to SAP GUI first, then re-run."
            )
        connection = app.Children(0)
        return connection.Children(0)

    def fetch_contracts(self, account_ids: List[str]) -> List[Account]:  # pragma: no cover
        """Run ZRGCSTAT and parse the exported contract list.

        Placeholder flow — confirm control IDs on the real screen:
            session.findById("wnd[0]/tbar[0]/okcd").text = "/nZRGCSTAT"
            session.findById("wnd[0]").sendVKey(0)
            # set Role Category = Buyers Only, execute, export to self.export_dir
        Then read the exported Excel and build Account rows.
        """
        session = self._session()
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nZRGCSTAT"
        session.findById("wnd[0]").sendVKey(0)
        # TODO: set "Buyers Only", execute (F8), export to Excel into export_dir.
        export_path = os.path.join(self.export_dir, "ZRGCSTAT_export.xlsx")
        if not os.path.exists(export_path):
            raise RuntimeError(
                f"Expected SAP export at {export_path} was not created. "
                "Confirm the ZRGCSTAT export step / field IDs."
            )
        return _parse_zrgcstat(export_path, account_ids)


def _parse_zrgcstat(path: str, account_ids: List[str]) -> List[Account]:  # pragma: no cover
    import openpyxl  # local import keeps base import light

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(max_row=1))]

    def col(*names: str) -> int:
        for n in names:
            if n in header:
                return header.index(n)
        return -1

    idx = {
        "account": col("account_id", "account"),
        "current": col("current_contract_no", "contract", "contract_no"),
        "prev1": col("previous_contract_no_1", "prev_contract_1"),
        "prev2": col("previous_contract_no_2", "prev_contract_2"),
        "plan": col("plan_type", "plan"),
    }
    wanted = set(account_ids)
    out: List[Account] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        acc = str(row[idx["account"]]) if idx["account"] >= 0 else ""
        if not acc or (wanted and acc not in wanted):
            continue
        out.append(
            Account(
                account_id=acc,
                client_name="",
                project="",
                unit="",
                current_contract_no=str(row[idx["current"]] or ""),
                previous_contract_no_1=_opt(row, idx["prev1"]),
                previous_contract_no_2=_opt(row, idx["prev2"]),
                plan_type=str(row[idx["plan"]] or "RF") if idx["plan"] >= 0 else "RF",
            )
        )
    return out


def _opt(row, i: int):  # pragma: no cover
    if i < 0:
        return None
    v = row[i]
    return str(v) if v not in (None, "") else None
